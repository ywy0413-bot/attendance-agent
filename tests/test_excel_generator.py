import pytest
from datetime import datetime, date, time
from io import BytesIO

from openpyxl import load_workbook

from src.report.excel_generator import ExcelReportGenerator
from src.models.attendance import AttendanceRecord
from src.models.vacation import VacationRecord
from src.email.email_classifier import AttendanceSubType


class TestExcelReportGenerator:
    """엑셀 보고서 생성기 테스트"""

    @pytest.fixture
    def generator(self):
        return ExcelReportGenerator()

    @pytest.fixture
    def sample_data(self):
        return {
            'vacations': [
                VacationRecord(
                    applicant="홍길동",
                    dates=[date(2024, 1, 15)],
                    department="개발팀",
                    vacation_type="연차",
                    reason="개인 사정",
                    email_received_at=datetime(2024, 1, 14, 10, 30)
                )
            ],
            'late_arrivals': [
                AttendanceRecord(
                    applicant="김철수",
                    sub_type=AttendanceSubType.LATE_ARRIVAL,
                    date=date(2024, 1, 15),
                    start_time=time(10, 0),
                    end_time=time(10, 30),
                    department="기획팀",
                    reason="교통 체증",
                    email_received_at=datetime(2024, 1, 15, 9, 0)
                )
            ],
            'outings': [
                AttendanceRecord(
                    applicant="이영희",
                    sub_type=AttendanceSubType.OUTING,
                    date=date(2024, 1, 15),
                    start_time=time(14, 0),
                    end_time=time(16, 0),
                    department="영업팀",
                    reason="거래처 미팅",
                    email_received_at=datetime(2024, 1, 15, 13, 0)
                )
            ],
            'early_leaves': [],
            'report_date': datetime(2024, 1, 15)
        }

    def test_generate_creates_valid_excel(self, generator, sample_data):
        """유효한 엑셀 파일 생성 테스트"""
        excel_bytes = generator.generate(sample_data)

        assert excel_bytes is not None
        assert len(excel_bytes) > 0

        # 엑셀 파일로 로드 가능한지 확인
        wb = load_workbook(BytesIO(excel_bytes))
        assert wb is not None

    def test_generate_has_correct_sheets(self, generator, sample_data):
        """올바른 시트 생성 테스트"""
        excel_bytes = generator.generate(sample_data)
        wb = load_workbook(BytesIO(excel_bytes))

        sheet_names = wb.sheetnames
        assert "요약" in sheet_names
        assert "휴가신고" in sheet_names
        assert "근태공유_출근지연" in sheet_names
        assert "근태공유_외출" in sheet_names
        assert "근태공유_조기퇴근" in sheet_names

    def test_summary_sheet_has_correct_counts(self, generator, sample_data):
        """요약 시트 건수 확인 테스트"""
        excel_bytes = generator.generate(sample_data)
        wb = load_workbook(BytesIO(excel_bytes))

        ws_summary = wb["요약"]

        # 건수 확인 (A4부터 데이터 시작)
        assert ws_summary['B4'].value == 1  # 휴가신고
        assert ws_summary['B5'].value == 1  # 출근지연
        assert ws_summary['B6'].value == 1  # 외출
        assert ws_summary['B7'].value == 0  # 조기퇴근
        assert ws_summary['B8'].value == 3  # 총계

    def test_vacation_sheet_has_data(self, generator, sample_data):
        """휴가신고 시트 데이터 확인 테스트"""
        excel_bytes = generator.generate(sample_data)
        wb = load_workbook(BytesIO(excel_bytes))

        ws_vacation = wb["휴가신고"]

        # 헤더 확인
        assert ws_vacation['A1'].value == "No"
        assert ws_vacation['B1'].value == "신청자"

        # 데이터 확인
        assert ws_vacation['A2'].value == 1
        assert ws_vacation['B2'].value == "홍길동"
        assert ws_vacation['C2'].value == "개발팀"
