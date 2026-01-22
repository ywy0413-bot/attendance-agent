"""임직원 한글 이름 → 영어 이름 변환 모듈"""
import logging
from pathlib import Path
from typing import Dict, Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class EmployeeMapper:
    """임직원 이름 매퍼"""

    def __init__(self, excel_path: str = None):
        """
        Args:
            excel_path: 임직원 정보 엑셀 파일 경로
        """
        if excel_path is None:
            # 기본 경로: 프로젝트 루트의 임직원 정보 파일
            project_root = Path(__file__).parent.parent.parent
            excel_files = list(project_root.glob("임직원*.xlsx"))
            if excel_files:
                excel_path = str(excel_files[0])
            else:
                logger.warning("임직원 정보 파일을 찾을 수 없습니다")
                self._name_map = {}
                return

        self._name_map = self._load_name_map(excel_path)
        logger.info(f"{len(self._name_map)}명의 임직원 정보를 로드했습니다")

    def _load_name_map(self, excel_path: str) -> Dict[str, str]:
        """엑셀 파일에서 한글→영어 이름 매핑 로드"""
        name_map = {}
        try:
            wb = load_workbook(excel_path, read_only=True)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 4:
                    korean_name = row[2]  # 3열: 성명
                    english_name = row[3]  # 4열: 영문호칭

                    if korean_name and english_name:
                        name_map[str(korean_name).strip()] = str(english_name).strip()

            wb.close()
        except Exception as e:
            logger.error(f"임직원 정보 로드 실패: {e}")

        return name_map

    def to_english(self, korean_name: str) -> str:
        """
        한글 이름을 영어 이름으로 변환

        Args:
            korean_name: 한글 이름

        Returns:
            영어 이름 (매핑 없으면 한글 이름 그대로 반환)
        """
        korean_name = korean_name.strip()
        return self._name_map.get(korean_name, korean_name)

    def get_all_mappings(self) -> Dict[str, str]:
        """모든 이름 매핑 반환"""
        return self._name_map.copy()


# 전역 인스턴스
_mapper = None


def get_employee_mapper() -> EmployeeMapper:
    """전역 EmployeeMapper 인스턴스 반환"""
    global _mapper
    if _mapper is None:
        _mapper = EmployeeMapper()
    return _mapper
