import io
import logging
from typing import List, Dict, Any
from datetime import datetime, date, time
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from models.attendance import AttendanceRecord
from models.vacation import VacationRecord
from utils.employee_mapper import get_employee_mapper

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """엑셀 보고서 생성기"""

    # 스타일 정의
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    SUMMARY_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
    LEFT_ALIGN = Alignment(horizontal='left', vertical='center')

    def __init__(self):
        self.employee_mapper = get_employee_mapper()

    def generate(self, data: Dict[str, Any], previous_deductions: Dict[str, List[Dict]] = None) -> bytes:
        """
        엑셀 보고서 생성 (근태공유 시트만)

        Args:
            data: 처리 결과 데이터
            previous_deductions: 이전 휴가차감 이력 (영문이름 -> [{'date': 날짜, 'minutes': 분}, ...])
        """
        if previous_deductions is None:
            previous_deductions = {}

        wb = Workbook()

        # 근태공유 시트 (출근지연, 외출, 조기퇴근 통합)
        ws_attendance = wb.active
        ws_attendance.title = "근태공유"
        self._create_attendance_combined_sheet(
            ws_attendance,
            data.get('late_arrivals', []),
            data.get('outings', []),
            data.get('early_leaves', []),
            previous_deductions
        )

        # 바이트로 저장
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        logger.info("엑셀 보고서가 생성되었습니다")
        return buffer.getvalue()

    def _create_summary_sheet(self, ws: Worksheet, data: Dict[str, Any]) -> None:
        """요약 시트 생성"""
        ws.title = "요약"

        # 보고서 제목
        report_date = data.get('report_date', datetime.now())
        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.value = f"휴가/근태 보고서 ({report_date.strftime('%Y-%m-%d')})"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = self.CENTER_ALIGN

        # 헤더
        headers = ["구분", "건수"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
            cell.alignment = self.CENTER_ALIGN

        # 데이터
        summary_data = [
            ("휴가신고", len(data.get('vacations', []))),
            ("출근지연", len(data.get('late_arrivals', []))),
            ("외출", len(data.get('outings', []))),
            ("조기퇴근", len(data.get('early_leaves', []))),
        ]

        total = sum(count for _, count in summary_data)

        for row_idx, (category, count) in enumerate(summary_data, 4):
            cell1 = ws.cell(row=row_idx, column=1, value=category)
            cell2 = ws.cell(row=row_idx, column=2, value=count)

            cell1.fill = self.SUMMARY_FILL
            cell1.border = self.BORDER
            cell1.alignment = self.CENTER_ALIGN

            cell2.border = self.BORDER
            cell2.alignment = self.CENTER_ALIGN

        # 총계
        total_row = len(summary_data) + 4
        total_cell1 = ws.cell(row=total_row, column=1, value="총계")
        total_cell2 = ws.cell(row=total_row, column=2, value=total)

        total_cell1.fill = self.TOTAL_FILL
        total_cell1.font = Font(bold=True)
        total_cell1.border = self.BORDER
        total_cell1.alignment = self.CENTER_ALIGN

        total_cell2.fill = self.TOTAL_FILL
        total_cell2.font = Font(bold=True)
        total_cell2.border = self.BORDER
        total_cell2.alignment = self.CENTER_ALIGN

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10

    def _create_vacation_sheet(self, ws: Worksheet, vacations: List[VacationRecord]) -> None:
        """휴가신고 시트 생성 (이름별 취합, 셀 병합)"""
        headers = ["No", "이름", "휴가일자", "휴가일수", "누적"]
        self._write_headers(ws, headers)

        # 이름별로 그룹화
        employee_vacations = defaultdict(list)
        for v in vacations:
            english_name = self.employee_mapper.to_english(v.applicant)
            employee_vacations[english_name].append(v)

        # 각 직원별로 행 생성
        current_row = 2
        employee_no = 1

        for name in sorted(employee_vacations.keys()):
            # 휴가일자 기준 오름차순 정렬
            vac_list = sorted(employee_vacations[name], key=lambda v: v.dates[0] if v.dates else date.max)
            start_row = current_row

            # 누적 계산
            total_days = sum(self._get_vacation_days(v) for v in vac_list)

            for i, v in enumerate(vac_list):
                # No, 이름은 첫 번째 행에만 작성
                if i == 0:
                    ws.cell(row=current_row, column=1, value=employee_no)
                    ws.cell(row=current_row, column=2, value=name)
                    ws.cell(row=current_row, column=5, value=total_days)

                # 휴가일자, 휴가일수는 모든 행에 작성
                ws.cell(row=current_row, column=3, value=v.date_range_str)
                ws.cell(row=current_row, column=4, value=self._get_vacation_days(v))

                # 스타일 적용 (모든 셀 가운데 정렬)
                for col in range(1, 6):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = self.BORDER
                    cell.alignment = self.CENTER_ALIGN

                current_row += 1

            # 셀 병합 (2개 이상 행일 때만)
            end_row = current_row - 1
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)  # No
                ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)  # 이름
                ws.merge_cells(start_row=start_row, start_column=5, end_row=end_row, end_column=5)  # 누적

            employee_no += 1

        self._auto_adjust_columns(ws)

    def _get_vacation_days(self, vacation: VacationRecord) -> float:
        """휴가일수 반환 (이메일에서 추출된 값 우선 사용)"""
        # 이메일에서 추출된 휴가일수가 있으면 사용
        if vacation.vacation_days is not None:
            return vacation.vacation_days

        # 없으면 휴가 종류에 따라 계산 (fallback)
        vtype = (vacation.vacation_type or "").lower()

        if '오전반차' in vtype or '오후반차' in vtype or '반반차' in vtype:
            return 0.25
        elif '반차' in vtype:
            return 0.5
        elif '연차' in vtype or '휴가' in vtype:
            return 1.0

        # 기본값
        return 1.0

    def _create_attendance_combined_sheet(
        self,
        ws: Worksheet,
        late_arrivals: List[AttendanceRecord],
        outings: List[AttendanceRecord],
        early_leaves: List[AttendanceRecord],
        previous_deductions: Dict[str, List[Dict]] = None
    ) -> None:
        """근태공유 통합 시트 생성 (셀 병합, 이전 차감 반영)"""
        if previous_deductions is None:
            previous_deductions = {}

        headers = [
            "이름",
            "출근지연-일자", "출근지연-시간(분)",
            "조기퇴근-일자", "조기퇴근-시간(분)",
            "외출-일자", "외출-시간(분)",
            "휴가차감-일자", "차감시간",
            "누계"
        ]
        self._write_headers(ws, headers)

        # 이름별로 데이터 그룹화
        employee_data = defaultdict(lambda: {
            'late_arrivals': [],
            'early_leaves': [],
            'outings': []
        })

        for record in late_arrivals:
            english_name = self.employee_mapper.to_english(record.applicant)
            employee_data[english_name]['late_arrivals'].append(record)

        for record in early_leaves:
            english_name = self.employee_mapper.to_english(record.applicant)
            employee_data[english_name]['early_leaves'].append(record)

        for record in outings:
            english_name = self.employee_mapper.to_english(record.applicant)
            employee_data[english_name]['outings'].append(record)

        # 각 직원별로 행 생성
        current_row = 2
        today_str = datetime.now().strftime("%Y-%m-%d")

        for name in sorted(employee_data.keys()):
            data = employee_data[name]
            start_row = current_row

            # 이전 차감 이력 조회 (날짜별 리스트)
            deduction_history = previous_deductions.get(name, [])
            already_deducted = sum(d['minutes'] for d in deduction_history)

            # 현재 폴더의 총 근태 시간 계산
            total_minutes = 0
            for r in data['late_arrivals']:
                total_minutes += self._calculate_minutes(r.start_time, r.end_time)
            for r in data['early_leaves']:
                total_minutes += self._calculate_minutes(r.start_time, r.end_time)
            for r in data['outings']:
                total_minutes += self._calculate_minutes(r.start_time, r.end_time)

            # 아직 차감되지 않은 시간 기준으로 새 차감 계산
            remaining_for_new_deduction = total_minutes - already_deducted
            new_deduction_days = self._calculate_deduction_days(remaining_for_new_deduction)
            new_deducted_minutes = int(new_deduction_days / 0.25) * 120 if new_deduction_days > 0 else 0

            # 오늘 새로 차감되는 경우 이력에 추가
            all_deductions = list(deduction_history)
            if new_deduction_days > 0:
                all_deductions.append({'date': today_str, 'minutes': new_deducted_minutes})

            # 총 차감 시간 (이전 + 신규)
            total_deducted = already_deducted + new_deducted_minutes

            # 누계 = 총 시간 - 총 차감 시간
            remaining_minutes = total_minutes - total_deducted

            # 최대 행 수 계산 (차감 이력 포함)
            max_rows = max(
                len(data['late_arrivals']),
                len(data['early_leaves']),
                len(data['outings']),
                len(all_deductions),
                1
            )

            for i in range(max_rows):
                # 이름, 누계는 첫 행에만
                if i == 0:
                    ws.cell(row=current_row, column=1, value=name)
                    ws.cell(row=current_row, column=10, value=remaining_minutes)

                # 휴가차감 이력 (날짜별)
                if i < len(all_deductions):
                    deduction = all_deductions[i]
                    ws.cell(row=current_row, column=8, value=deduction['date'])
                    ws.cell(row=current_row, column=9, value=deduction['minutes'])

                # 출근지연
                if i < len(data['late_arrivals']):
                    r = data['late_arrivals'][i]
                    ws.cell(row=current_row, column=2, value=r.date.strftime("%Y-%m-%d") if r.date else "")
                    ws.cell(row=current_row, column=3, value=self._calculate_minutes(r.start_time, r.end_time))

                # 조기퇴근
                if i < len(data['early_leaves']):
                    r = data['early_leaves'][i]
                    ws.cell(row=current_row, column=4, value=r.date.strftime("%Y-%m-%d") if r.date else "")
                    ws.cell(row=current_row, column=5, value=self._calculate_minutes(r.start_time, r.end_time))

                # 외출
                if i < len(data['outings']):
                    r = data['outings'][i]
                    ws.cell(row=current_row, column=6, value=r.date.strftime("%Y-%m-%d") if r.date else "")
                    ws.cell(row=current_row, column=7, value=self._calculate_minutes(r.start_time, r.end_time))

                # 스타일 적용 (모든 셀 가운데 정렬)
                for col in range(1, 11):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = self.BORDER
                    cell.alignment = self.CENTER_ALIGN

                current_row += 1

            # 셀 병합 (2개 이상 행일 때만) - 이름과 누계만 병합
            end_row = current_row - 1
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)  # 이름
                ws.merge_cells(start_row=start_row, start_column=10, end_row=end_row, end_column=10)  # 누계

        self._auto_adjust_columns(ws)

    def _calculate_deduction_days(self, total_minutes: int) -> float:
        """누적 분을 휴가차감 일수로 변환 (120분 단위)"""
        if total_minutes < 120:
            return 0.0
        return (total_minutes // 120) * 0.25

    def _calculate_minutes(self, start_time, end_time) -> int:
        """시간 차이를 분으로 계산"""
        if not start_time or not end_time:
            return 0
        start_minutes = start_time.hour * 60 + start_time.minute
        end_minutes = end_time.hour * 60 + end_time.minute
        return max(0, end_minutes - start_minutes)

    def _write_headers(self, ws: Worksheet, headers: List[str]) -> None:
        """헤더 행 작성"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
            cell.alignment = self.CENTER_ALIGN

    def _write_row(self, ws: Worksheet, row: int, data: List[Any]) -> None:
        """데이터 행 작성 (모든 셀 가운데 정렬)"""
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = self.BORDER
            cell.alignment = self.CENTER_ALIGN

    def _auto_adjust_columns(self, ws: Worksheet) -> None:
        """컬럼 너비 자동 조정"""
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    cell_value = str(cell.value) if cell.value else ""
                    cell_length = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            adjusted_width = max(min(max_length + 2, 50), 10)
            ws.column_dimensions[column].width = adjusted_width
