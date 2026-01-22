"""임직원 정보 파일 구조 확인"""
from openpyxl import load_workbook

wb = load_workbook("임직원 정보_Rev.0_Lara_260112.xlsx")
ws = wb.active

print(f"시트 이름: {ws.title}")
print(f"열 헤더 (1행):")
for col in range(1, 10):
    cell = ws.cell(row=1, column=col)
    print(f"  {col}열: {cell.value}")

print(f"\n샘플 데이터 (2-5행):")
for row in range(2, 6):
    row_data = []
    for col in range(1, 10):
        row_data.append(str(ws.cell(row=row, column=col).value or ""))
    print(f"  {row}행: {row_data}")
